from bs4 import BeautifulSoup  # импортируем библиотеку BeautifulSoup
import requests  # импортируем библиотеку requests
from openpyxl import load_workbook


def parse1():
    fn = 'лаба1.xlsx'
    wb = load_workbook(fn)
    ws = wb['данные']
    description1 = []  # Информация оОООООО
    description2 = []
    description3 = []
    for i in range(1, 2 + 1):
        url1 = 'https://www.chitai-gorod.ru/catalog/collections/bestsell?page=' + str(i)  # передаем необходимы URL адрес
        page1 = requests.get(url1)  # отправляем запрос методом Get на данный адрес и получаем ответ в переменную
        print(page1.status_code)  # смотрим ответ
        soup1 = BeautifulSoup(page1.text, "html.parser")  # передаем страницу в bs4

        block1 = soup1.findAll('article', class_='product-card product-card product')  # находим контейнер с нужным классом

        for data in block1:  # проходим циклом по содержимому контейнера
            nazv = data.find(class_='product-title__head')
            if nazv is not None:
                description1.append(nazv.text)

            author = data.find(class_='product-title__author')
            if author is not None:
                description2.append(author.text)

            price = data.find(class_='product-price__value')
            if price is not None:
                description3.append(price.text)

    print(description1, description2, description3)
    ech = 1
    for elem1, elem2, elem3 in zip(description1, description2, description3):
        cell = ws.cell(1, ech)
        cell.value = elem1

        cell = ws.cell(2, ech)
        cell.value = elem2

        cell = ws.cell(3, ech)
        cell.value = elem3

        ech += 1

    wb.save(fn)
    wb.close()


if __name__ == '__main__':
    parse1()
